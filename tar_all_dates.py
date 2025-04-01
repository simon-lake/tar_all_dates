import time
import pandas as pd
from scipy.stats import linregress, pearsonr
import glob
import matplotlib.pyplot as plt
import os
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
import shutil


def process_planting_date(planting_date_dir, base_output_dir):
    """
    Processes data for a single planting date.

    Args:
        planting_date_dir (str): The directory for the planting date (e.g., '01a010').
        base_output_dir (str): The base output directory for all planting dates.
    """

    start_time_section1 = time.time()

    # Construct input and output directories
    csv_directory = os.path.join(base_input_dir, planting_date_dir, 'input')
    output_directory = os.path.join(
        base_output_dir, planting_date_dir, 'output')

    # Create output directory if it doesn't exist
    os.makedirs(output_directory, exist_ok=True)

    # Use glob to get the list of CSV files in the specified directory
    csv_files = glob.glob(os.path.join(csv_directory, '*.csv'))

    # List of files you want to include
    files_to_include = [
        'obstmp_195099_tardry_' + planting_date_dir + '_wue_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_hin_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_bac_pcv_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_rfl_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_yld_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_yld_pcv_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_sst_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_ste_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_sle_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_evp_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_tra_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_trm_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_erm_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_bmx_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_bmx_pcv_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_len_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_etc_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_wue_pcv_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_cfa_sum_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_yld_obs_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_map_ave_014.csv',
        'obstmp_195099_tardry_' + planting_date_dir + '_bac_ave_014.csv'
    ]

    # Filter the files based on the list of files to include
    filtered_csv_files = [
        file for file in csv_files if os.path.basename(file) in files_to_include]

    # Create an empty DataFrame to store the combined data
    combined_data = pd.DataFrame()

    # Iterate through the file paths and load the data
    for file_path in filtered_csv_files:
        # Read the first column from the first file and both columns from the rest
        if file_path == filtered_csv_files[0]:
            df = pd.read_csv(file_path)
        else:
            df = pd.read_csv(file_path, usecols=[1])

        # Concatenate the data to the combined DataFrame
        combined_data = pd.concat([combined_data, df], axis=1)

    # Calculate ROCF based on the logic provided in the rcf_ave code
    rcf_ave = []
    cfa_sum = combined_data.iloc[:, 5]
    yld_obs = combined_data.iloc[:, 21]

    for cfa, yld in zip(cfa_sum, yld_obs):
        if cfa == -999:
            rcf_ave.append(-999)
        else:
            if yld <= 48:
                adjusted_cfa = cfa + (49 - yld)
            else:
                adjusted_cfa = cfa
            rcf_value = (adjusted_cfa / 49) * 100
            rcf_ave.append(rcf_value)

    # Add the ROCF column to the combined DataFrame
    combined_data['RCF'] = rcf_ave

    # Rename the columns in row 0
    combined_data.columns = ['sub_cat', 'BAC (t ha^-1)', 'BAC_CV (%)', 'BMX (%)', 'BMX_CV (%)', 'CFA (sum)', 'Es/Esm (%)', 'ETC (%)', 'EVP (mm)', 'Harvest index (%)', 'CROP CYCLE (days)', 'MAP (mm)', 'RFL (mm)', 'LEAF STRESS (%)', 'STO. STRESS (%)',
                             'TEMP. STRESS (%)', 'TRA (mm)', 'ET/ETm (%)', 'CWP (kg m^-3)', 'CWP CV (%)', 'YIELD (t ha^-1)', 'YLD (obs)', 'YIELD CV (%)', 'RCF (%)']

    # Write the combined data to a new Excel file in the output directory
    combined_data_output_path = os.path.join(
        output_directory, 'combined_data.xlsx')
    combined_data.to_excel(combined_data_output_path, index=False)

    # Print a message indicating that the operation is done
    print(
        f"Data combination completed for {planting_date_dir}. The combined data is saved to 'combined_data.xlsx'.")

    # Filter out rows with any column containing -999
    combined_data_filtered = combined_data[(combined_data != -999).all(axis=1)]

    # Remove columns 2 (cfa sum) from the dataframe
    combined_data_filtered = combined_data_filtered.drop(
        columns=[combined_data.columns[5]])

    # Write the filtered data to a new Excel file in the output directory
    filtered_data_output_path = os.path.join(
        output_directory, 'combined_data_filtered.xlsx')
    combined_data_filtered.to_excel(filtered_data_output_path, index=False)

    # Print a message indicating that the operation is done
    print(
        f"Data filtering completed for {planting_date_dir}. The combined data is saved to 'combined_data_filtered.xlsx'.")

    # Specify the full path to the Excel file
    excel_file_path = filtered_data_output_path

    # Load the Excel file
    data = pd.read_excel(excel_file_path)

    # Get the column names
    columns = data.columns

    # Create a dictionary to store the R-squared values
    statistics = {}

    # Iterate through each pair of columns and calculate R-squared
    for i in range(1, len(columns)):
        for j in range(i + 1, len(columns)):
            x = data[columns[i]]
            y = data[columns[j]]

            # Check if any of the variables have all 0 values
            if (x == 0).all() or (y == 0).all():
                # Handle the case where one of the variables is all 0
                r_squared = None
                pearson_corr = None
            else:
                # Perform linear regression
                slope, intercept, r_value, p_value, std_err = linregress(x, y)

                # Calculate R-squared
                r_squared = r_value**2

                # Calculate Pearson correlation coefficient
                pearson_corr, _ = pearsonr(x, y)
                pearson_corr = abs(pearson_corr)

            # Store the R-squared and Pearson correlation coefficient values in the dictionary
            key = f"{columns[i]} vs {columns[j]}"
            statistics[key] = {
                'R-squared': r_squared,
                'Pearson Correlation': pearson_corr
            }

            # Check if Pearson correlation coefficient is greater than 0.8, then plot a scatter chart
            if pearson_corr is not None and pearson_corr > 0.8:
                plt.scatter(x, y)
                plt.xlabel(columns[i])
                plt.ylabel(columns[j])

                # Plot the linear regression line
                plt.plot(x, slope * x + intercept, color='black', linestyle='dotted',
                         label=f'y = {slope:.2f}x + {intercept:.2f}, R2: {r_squared:.2f}')

                # Display legend
                legend_text = (f'Equation: y = {slope:.2f}x + {intercept:.2f}\n'
                               f'R²: {r_squared:.2f}, Pearson r: {pearson_corr:.2f}')
                plt.legend([legend_text], loc='upper left', handlelength=0)
                plt.grid()

                plt.ylim(bottom=0)

                # Save the PNG file with the relationship name in the output directory
                relationship_name = f"{columns[i]}_vs_{columns[j]}".replace(
                    '/', '_').replace(' ', '_')
                file_path = os.path.join(
                    output_directory, f"{relationship_name}_scatter_plot.png")
                plt.savefig(file_path)

                # Close the current figure to avoid overlap with the next plot
                plt.close()

    # Create a dataframe from the R-squared and Pearson correlation coefficient values
    df_r_squared = pd.DataFrame.from_dict(statistics, orient='index')

    # Save the results to an Excel file
    output_filename = "r_squared_values_with_equation.xlsx"
    output_filepath = os.path.join(output_directory, output_filename)
    df_r_squared.to_excel(output_filepath, index=True)

    # Open the Excel file using openpyxl
    wb = openpyxl.load_workbook(output_filepath)
    ws = wb.active

    # Iterate through each row and check the Pearson correlation value
    # Start from row 2 (header is in row 1)
    for row_idx, row in enumerate(df_r_squared.itertuples(), start=2):
        pearson_corr = row._2

        # Check if Pearson correlation is greater than 0.8
        if pearson_corr is not None and pearson_corr > 0.8:
            # Highlight the entire row in yellow
            for col_idx in range(1, len(df_r_squared.columns) + 2):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Save the modified Excel file
    wb.save(output_filepath)

    # Print a message indicating that the operation is done
    print(
        f"Data stats completed for {planting_date_dir}. The correlation stats are saved to 'r_squared_values_with_equation.xlsx'.")

    # Print the column names for debugging purposes
    print("Column names in df_r_squared:", df_r_squared.columns)

    time.sleep(1)  # Simulating process
    end_time_section1 = time.time()
    print(
        f"Section 1 for {planting_date_dir} took {end_time_section1 - start_time_section1:.2f} seconds.")

    # --- Section 2 ---
    start_time_section2 = time.time()

    # --- BMX and HIN NOM ---
    def calculate_nom(file_prefix, nom_name, csv_directory, planting_date_dir):
        """
        Calculates the normalized value (nom) for a given file.

        Args:
            file_prefix (str): The prefix of the file name (e.g., 'bmx_ave', 'hin_ave').
            nom_name (str): The name of the new normalized column (e.g., 'bmx_nom', 'hin_nom').
            csv_directory (str): The directory containing the CSV files.
            planting_date_dir (str): The planting date directory (e.g., '01j010').
        """
        file_name = f'obstmp_195099_tardry_{planting_date_dir}_{file_prefix}_014.csv'
        file_path = os.path.join(csv_directory, file_name)
        try:
            data = pd.read_csv(file_path)
            max_value = data.iloc[:, 1].max()
            data[nom_name] = (data.iloc[:, 1] / max_value) * 100
            nom_data = data[['sub_cat', nom_name]]
            nom_output_path = os.path.join(
                csv_directory, file_name.replace('_ave_', '_nom_'))
            nom_data.to_csv(nom_output_path, index=False)
            print(
                f"{nom_name.upper()} file created and saved to '{nom_output_path}'.")
        except FileNotFoundError:
            print(
                f"Warning: {file_name} not found for {planting_date_dir}. Skipping.")
        except Exception as e:
            print(
                f"An error occurred while processing {file_name} for {planting_date_dir}: {e}")

    calculate_nom('bmx_ave', 'bmx_nom', csv_directory, planting_date_dir)
    calculate_nom('hin_ave', 'hin_nom', csv_directory, planting_date_dir)

    # --- NWP ---
    wue_ave_file_path = os.path.join(
        csv_directory, f'obstmp_195099_tardry_{planting_date_dir}_wue_ave_014.csv')
    try:
        wue_ave_data = pd.read_csv(wue_ave_file_path)
        nwp_k = []
        nwp_fe = []
        nwp_zn = []

        for wue in wue_ave_data.iloc[:, 1]:
            if wue == -999:
                nwp_k.append(-999)
                nwp_fe.append(-999)
                nwp_zn.append(-999)
            else:
                nwp_k.append(wue * 19.12)
                nwp_fe.append(wue * 0.284)
                nwp_zn.append(wue * 0.137)

        nwp_k_data = pd.DataFrame(
            {'sub_cat': wue_ave_data.iloc[:, 0], 'nwp_k': nwp_k})
        nwp_fe_data = pd.DataFrame(
            {'sub_cat': wue_ave_data.iloc[:, 0], 'nwp_fe': nwp_fe})
        nwp_zn_data = pd.DataFrame(
            {'sub_cat': wue_ave_data.iloc[:, 0], 'nwp_zn': nwp_zn})

        nwp_k_output_path = os.path.join(
            csv_directory, f'obstmp_195099_tardry_{planting_date_dir}_nwp_k_ave_014.csv')
        nwp_fe_output_path = os.path.join(
            csv_directory, f'obstmp_195099_tardry_{planting_date_dir}_nwp_fe_ave_014.csv')
        nwp_zn_output_path = os.path.join(
            csv_directory, f'obstmp_195099_tardry_{planting_date_dir}_nwp_zn_ave_014.csv')

        nwp_k_data.to_csv(nwp_k_output_path, index=False)
        nwp_fe_data.to_csv(nwp_fe_output_path, index=False)
        nwp_zn_data.to_csv(nwp_zn_output_path, index=False)

        print(f"nwp_k file created and saved to '{nwp_k_output_path}'.")
        print(f"nwp_fe file created and saved to '{nwp_fe_output_path}'.")
        print(f"nwp_zn file created and saved to '{nwp_zn_output_path}'.")
    except FileNotFoundError:
        print(
            f"Warning: wue_ave file not found for {planting_date_dir}. Skipping NWP calculations.")
    except Exception as e:
        print(
            f"An error occurred while processing wue_ave for {planting_date_dir}: {e}")

    # --- RCF ---
    cfa_sum_file_path = os.path.join(
        csv_directory, f'obstmp_195099_tardry_{planting_date_dir}_cfa_sum_014.csv')
    yld_obs_file_path = os.path.join(
        csv_directory, f'obstmp_195099_tardry_{planting_date_dir}_yld_obs_014.csv')
    try:
        cfa_sum_data = pd.read_csv(cfa_sum_file_path)
        yld_obs_data = pd.read_csv(yld_obs_file_path)

        rcf_ave = []
        for cfa, yld in zip(cfa_sum_data.iloc[:, 1], yld_obs_data.iloc[:, 1]):
            if cfa == -999:
                rcf_ave.append(-999)
            else:
                if yld <= 48:
                    adjusted_cfa = cfa + (49 - yld)
                else:
                    adjusted_cfa = cfa
                rcf_value = (adjusted_cfa / 49) * 100
                rcf_ave.append(rcf_value)

        rcf_ave_data = pd.DataFrame(
            {'sub_cat': cfa_sum_data.iloc[:, 0], 'rcf_ave': rcf_ave})
        rcf_ave_output_path = os.path.join(
            csv_directory, f'obstmp_195099_tardry_{planting_date_dir}_rcf_ave_014.csv')
        rcf_ave_data.to_csv(rcf_ave_output_path, index=False)
        print(f"RCF_AVE file created and saved to '{rcf_ave_output_path}'.")
    except FileNotFoundError:
        print(
            f"Warning: cfa_sum or yld_obs file not found for {planting_date_dir}. Skipping RCF calculation.")
    except Exception as e:
        print(
            f"An error occurred while processing cfa_sum or yld_obs for {planting_date_dir}: {e}")

    # --- Suitability ---
    selected_csv_files = [f'obstmp_195099_tardry_{planting_date_dir}_bmx_ave_014.csv', f'obstmp_195099_tardry_{planting_date_dir}_wue_pcv_014.csv',
                          f'obstmp_195099_tardry_{planting_date_dir}_rfl_ave_014.csv', f'obstmp_195099_tardry_{planting_date_dir}_rcf_ave_014.csv']

    csv_conditions = {
        f'obstmp_195099_tardry_{planting_date_dir}_wue_pcv_014.csv': {'condition': lambda x: 0 if x == -999 else (-1 if x > 150 else 1), 'column_name': 'wue_cv_result'},
        f'obstmp_195099_tardry_{planting_date_dir}_rfl_ave_014.csv': {'condition': lambda x: 0 if x == -999 else (-1 if x < 200 else 1), 'column_name': 'rfl_result'},
        f'obstmp_195099_tardry_{planting_date_dir}_rcf_ave_014.csv': {'condition': lambda x: 0 if x == -999 else (-1 if x > 33 else 1), 'column_name': 'rcf_result'}
    }

    result_df = pd.DataFrame()
    first_column_added = False

    for csv_file in selected_csv_files:
        if csv_file in csv_conditions:
            file_path = os.path.join(csv_directory, csv_file)
            try:
                df = pd.read_csv(file_path)
                if not first_column_added:
                    result_df['sub_cat'] = df.iloc[:, 0]
                    first_column_added = True
                result_df[csv_conditions[csv_file]['column_name']] = np.vectorize(
                    csv_conditions[csv_file]['condition'])(df.iloc[:, 1])
            except FileNotFoundError:
                print(
                    f"Warning: {csv_file} not found for {planting_date_dir}. Skipping.")
            except Exception as e:
                print(
                    f"An error occurred while processing {csv_file} for {planting_date_dir}: {e}")

    if not result_df.empty:
        result_df['elim_result'] = result_df.iloc[:, 1:].apply(
            lambda row: -1 if (-1 in row.values) and (0 not in row.values) else (0 if 0 in row.values else (
                1 if all(x == 1 for x in row.values) else None)),
            axis=1
        )
        result_df['elimination'] = np.where(result_df['elim_result'] == 0, 0,
                                            np.where(result_df['elim_result'] == -1, 1, 2))
        suitability_output_path = os.path.join(
            output_directory, f'suitability_{planting_date_dir}.csv')
        result_df.to_csv(suitability_output_path, index=False)
        print(
            f"Suitability file created and saved to '{suitability_output_path}'.")
    else:
        print(f"No suitability data to process for {planting_date_dir}")

    time.sleep(2)  # Simulating process
    end_time_section2 = time.time()
    print(
        f"Section 2 for {planting_date_dir} took {end_time_section2 - start_time_section2:.2f} seconds.")

    # --- Section 3 ---
    start_time_section3 = time.time()

    def calculate_suitability(file_prefix, nom_name, csv_directory, planting_date_dir, output_directory, result_df):
        """
        Calculates suitability based on a given file.

        Args:
            file_prefix (str): The prefix of the file name (e.g., 'bmx_ave', 'hin_ave').
            nom_name (str): The name of the new normalized column (e.g., 'bmx_nom', 'hin_nom').
            csv_directory (str): The directory containing the CSV files.
            planting_date_dir (str): The planting date directory (e.g., '01j010').
            output_directory (str): The output directory.
            result_df (pd.DataFrame): The DataFrame to merge with and perform calculations on.
        Returns:
            pd.DataFrame: The updated result_df.
        """
        file_name = f'obstmp_195099_tardry_{planting_date_dir}_{file_prefix}_014.csv'
        file_path = os.path.join(csv_directory, file_name)
        try:
            # Read only column 1 from the CSV file
            column1 = pd.read_csv(file_path, usecols=[1], names=[
                                  nom_name.replace('_nom', '')])

            # Convert column to numeric, replacing non-numeric values with NaN
            column1[nom_name.replace('_nom', '')] = pd.to_numeric(
                column1[nom_name.replace('_nom', '')], errors='coerce')

            # Merge the DataFrame with the result DataFrame based on the 'sub_cat' column
            result_df = pd.merge(result_df, column1,
                                 left_on='sub_cat', right_index=True, how='left')

            # Apply the condition to set values to NaN based on the 'elimination' column
            result_df[nom_name.replace('_nom', '')] = np.where((result_df['elimination'] == 0) | (
                result_df['elimination'] == 1), np.nan, result_df[nom_name.replace('_nom', '')])

            # Calculate the maximum value for the column
            max_value = result_df[nom_name.replace('_nom', '')].max()

            # Normalize by dividing each value by the max value and then multiply by 100
            result_df[nom_name] = (
                result_df[nom_name.replace('_nom', '')] / max_value) * 100

            # Calculate the suitability based on the column
            result_df['S4'] = np.where(result_df[nom_name] < 25, 4, 0)
            result_df['S3'] = np.where(
                (result_df[nom_name] >= 25) & (result_df[nom_name] <= 50), 3, 0)
            result_df['S2'] = np.where(
                (result_df[nom_name] > 50) & (result_df[nom_name] <= 75), 2, 0)
            result_df['S1'] = np.where(result_df[nom_name] > 75, 1, 0)

            # Replace NaN values with 0 in the new columns
            result_df[['S4', 'S3', 'S2', 'S1']] = result_df[[
                'S4', 'S3', 'S2', 'S1']].fillna(0)

            # Create the 'final_suitability' column by summing the suitability columns
            result_df['final_suitability'] = result_df['S4'] + \
                result_df['S3'] + result_df['S2'] + result_df['S1']

            # Write the result DataFrame to an Excel file
            output_path = os.path.join(
                output_directory, f'suitability_analysis_{file_prefix}.xlsx')
            result_df.to_excel(output_path, index=False)

            # Define the output path for the CSV file
            csv_output_path = os.path.join(
                output_directory, f'suitability_analysis_{file_prefix}_all.csv')

            # Write the result DataFrame to a CSV file
            result_df.to_csv(csv_output_path, index=False)

            print(f"Results written to {output_path} and {csv_output_path}")

            # Read the comprehensive CSV file
            df = pd.read_csv(csv_output_path)

            # Keep only the desired columns
            columns_to_keep = ['sub_cat', 'elimination', 'final_suitability']
            df = df[columns_to_keep]

            # Define the output path for the modified CSV file
            modified_csv_output_path = os.path.join(
                output_directory, f'suitability_analysis_{file_prefix}.csv')

            # Write the modified DataFrame to a new CSV file
            df.to_csv(modified_csv_output_path, index=False)
            print(f"Modified file saved to {modified_csv_output_path}")

            # Replace the suitability levels with custom labels
            result_df['S4'] = result_df['S4'].replace({4: 'S4'})
            result_df['S3'] = result_df['S3'].replace({3: 'S3'})
            result_df['S2'] = result_df['S2'].replace({2: 'S2'})
            result_df['S1'] = result_df['S1'].replace({1: 'S1'})

            # Exclude zeros from the data
            non_zero_data = result_df[[
                'S4', 'S3', 'S2', 'S1']].replace(0, np.nan)

            # Count the occurrences of each suitability level
            counts = non_zero_data.apply(lambda col: col.value_counts()).T

            # Plot the bar chart
            ax = counts.plot(kind='bar', color=[
                             'blue', 'blue', 'blue'], stacked=True, legend=False)

            # Set labels and title
            ax.set_xlabel('Suitability_classes')
            ax.set_ylabel('Frequency')

            # Add data labels to the top of each bar
            for p in ax.patches:
                height = p.get_height()
                if height > 0:  # Exclude 0 counts
                    ax.annotate(f'{height:.0f}', (p.get_x() + p.get_width() / 2., height),
                                ha='center', va='center', xytext=(0, 5), textcoords='offset points')

            # Customize x-axis tick labels for S3, S2, and S1 with rotation
            ax.set_xticklabels(['S4', 'S3', 'S2', 'S1'],
                               rotation=0, ha='right')

            # Save the plot as a PNG file
            output_filepath = os.path.join(
                output_directory, f'suitability_analysis_{file_prefix}_plot.png')
            plt.savefig(output_filepath)
            plt.close()
            return result_df

        except FileNotFoundError:
            print(
                f"Warning: {file_name} not found for {planting_date_dir}. Skipping.")
            return result_df
        except Exception as e:
            print(
                f"An error occurred while processing {file_name} for {planting_date_dir}: {e}")
            return result_df

    def section_3(csv_directory, planting_date_dir, output_directory, result_df):
        """
        Performs Section 3 calculations for a given planting date.

        Args:
            csv_directory (str): The directory containing the CSV files.
            planting_date_dir (str): The planting date directory (e.g., '01j010').
            output_directory (str): The output directory.
            result_df (pd.DataFrame): The DataFrame to merge with and perform calculations on.
        """

        result_df = calculate_suitability(
            'bmx_ave', 'bmx_nom', csv_directory, planting_date_dir, output_directory, result_df)
        result_df = calculate_suitability(
            'hin_ave', 'hin_nom', csv_directory, planting_date_dir, output_directory, result_df)

    # Initialize result_df for section 3
    selected_csv_files = [f'obstmp_195099_tardry_{planting_date_dir}_hin_ave_014.csv', f'obstmp_195099_tardry_{planting_date_dir}_wue_pcv_014.csv',
                          f'obstmp_195099_tardry_{planting_date_dir}_rfl_ave_014.csv', f'obstmp_195099_tardry_{planting_date_dir}_rcf_ave_014.csv']

    csv_conditions = {
        f'obstmp_195099_tardry_{planting_date_dir}_wue_pcv_014.csv': {'condition': lambda x: 0 if x == -999 else (-1 if x > 150 else 1), 'column_name': 'wue_cv_result'},
        f'obstmp_195099_tardry_{planting_date_dir}_rfl_ave_014.csv': {'condition': lambda x: 0 if x == -999 else (-1 if x < 200 else 1), 'column_name': 'rfl_result'},
        f'obstmp_195099_tardry_{planting_date_dir}_rcf_ave_014.csv': {'condition': lambda x: 0 if x == -999 else (-1 if x > 33 else 1), 'column_name': 'rcf_result'}
    }

    result_df = pd.DataFrame()
    first_column_added = False

    for csv_file in selected_csv_files:
        if csv_file in csv_conditions:
            file_path = os.path.join(csv_directory, csv_file)
            try:
                df = pd.read_csv(file_path)
                if not first_column_added:
                    result_df['sub_cat'] = df.iloc[:, 0]
                    first_column_added = True
                result_df[csv_conditions[csv_file]['column_name']] = np.vectorize(
                    csv_conditions[csv_file]['condition'])(df.iloc[:, 1])
            except FileNotFoundError:
                print(
                    f"Warning: {csv_file} not found for {planting_date_dir}. Skipping.")
            except Exception as e:
                print(
                    f"An error occurred while processing {csv_file} for {planting_date_dir}: {e}")

    if not result_df.empty:
        result_df['elim_result'] = result_df.iloc[:, 1:].apply(
            lambda row: -1 if (-1 in row.values) and (0 not in row.values) else (0 if 0 in row.values else (
                1 if all(x == 1 for x in row.values) else None)),
            axis=1
        )
        result_df['elimination'] = np.where(result_df['elim_result'] == 0, 0,
                                            np.where(result_df['elim_result'] == -1, 1, 2))
    else:
        print(f"No suitability data to process for {planting_date_dir}")

    section_3(csv_directory, planting_date_dir, output_directory, result_df)

    end_time_section3 = time.time()
    print(
        f"Section 3 for {planting_date_dir} took {end_time_section3 - start_time_section3:.2f} seconds.")

### Section 4###


def section_4(base_input_dir, planting_date_dir, base_output_dir):
    """
    Performs Section 4 calculations for a given planting date.

    Args:
        base_input_dir (str): The base input directory.
        planting_date_dir (str): The planting date directory (e.g., '01a010').
        base_output_dir (str): The base output directory.
    """
    start_time_section4 = time.time()

    # Source directory containing the CSV files
    csv_directory = os.path.join(base_input_dir, planting_date_dir, 'input')
    output_directory = os.path.join(
        base_output_dir, planting_date_dir, 'output')

    # Dictionary mapping filename parts to target directories
    file_mappings = {
        'bmx_ave': os.path.join(base_output_dir, planting_date_dir, 'BMX'),
        'len_ave': os.path.join(base_output_dir, planting_date_dir, 'crop_cycle'),
        'bmx_nom': os.path.join(base_output_dir, planting_date_dir, 'BMX_nom'),
        'yld_ave': os.path.join(base_output_dir, planting_date_dir, 'yield'),
        'yld_pcv': os.path.join(base_output_dir, planting_date_dir, 'yield_cv'),
        'wue_ave': os.path.join(base_output_dir, planting_date_dir, 'CWP'),
        'rfl_ave': os.path.join(base_output_dir, planting_date_dir, 'RFL'),
        'bmx_pcv': os.path.join(base_output_dir, planting_date_dir, 'BMX_cv'),
        'yld_obs': os.path.join(base_output_dir, planting_date_dir, 'yield_obs'),
        'ste_ave': os.path.join(base_output_dir, planting_date_dir, 'temp_stress'),
        'sst_ave': os.path.join(base_output_dir, planting_date_dir, 'sto_stress'),
        'sle_ave': os.path.join(base_output_dir, planting_date_dir, 'leaf_stress'),
        'rcf_ave': os.path.join(base_output_dir, planting_date_dir, 'RCF'),
        'hin_ave': os.path.join(base_output_dir, planting_date_dir, 'HIN'),
        'hin_nom': os.path.join(base_output_dir, planting_date_dir, 'HIN_nom'),
        'nwp_k_ave': os.path.join(base_output_dir, planting_date_dir, 'nwp_k'),
        'nwp_fe_ave': os.path.join(base_output_dir, planting_date_dir, 'nwp_fe'),
        'nwp_zn_ave': os.path.join(base_output_dir, planting_date_dir, 'nwp_zn'),
    }

    # Get the list of all CSV files in the source directory
    csv_files = glob.glob(os.path.join(csv_directory, '*.csv'))

    # Iterate through the CSV files and copy them to their respective target directories
    for csv_file in csv_files:
        # Get the base name of the file (e.g., 'obstmp_195099_tardry_01a010_bmx_ave_014.csv')
        base_name = os.path.basename(csv_file)

        # Check which part of the filename matches the keys in the dictionary
        for key, target_directory in file_mappings.items():
            if key in base_name:
                # Ensure the target directory exists
                if not os.path.exists(target_directory):
                    os.makedirs(target_directory)

                # Specify the full path for the copied file
                destination_file_path = os.path.join(
                    target_directory, base_name)

                # Copy the file to the target directory
                shutil.copy(csv_file, destination_file_path)

                # Print a message indicating that the file has been copied
                print(f"File '{base_name}' copied to '{target_directory}'.")

    def process_data(csv_path, excel_path, csv_output_path, value_ranges, sheet_name):
        # Read the CSV file
        try:
            df = pd.read_csv(csv_path)
        except FileNotFoundError:
            print(f"Warning: {csv_path} not found. Skipping.")
            return
        except pd.errors.EmptyDataError:
            print(f"Warning: {csv_path} is empty. Skipping.")
            return
        except Exception as e:
            print(f"An error occurred while reading {csv_path}: {e}")
            return

        # Create a new DataFrame with the desired structure
        processed_df = pd.DataFrame()
        processed_df['sub_cat'] = df.iloc[:, 0]
        processed_df[sheet_name] = df.iloc[:, 1]

        # Add new columns with conditional logic based on value ranges
        for label, (min_val, max_val, output) in value_ranges.items():
            processed_df[label] = processed_df[sheet_name].apply(
                lambda x: output if (
                    x == -999) else (output if min_val < x <= max_val else 0) if x != -999 else 0
            )

        # Set other columns to 0 if -999 was detected in the first column
        processed_df['-999'] = processed_df[sheet_name].apply(
            lambda x: -1 if x == -999 else 0)

        for label in value_ranges.keys():
            if label != '-999':
                processed_df[label] = processed_df.apply(
                    lambda row: 0 if row['-999'] == -1 else row[label], axis=1
                )

        # Add the final column that sums the conditionally created columns
        processed_df['Final'] = processed_df.iloc[:,
                                                  2:len(value_ranges) + 2].sum(axis=1)

        # Write the DataFrame to an Excel file
        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                processed_df.to_excel(writer, index=False,
                                      header=True, sheet_name=sheet_name)
        except Exception as e:
            print(f"An error occurred while writing to {excel_path}: {e}")
            return

        # Drop columns that are not needed in the CSV output
        processed_df_csv = processed_df.drop(
            processed_df.columns[2:len(value_ranges) + 2], axis=1)

        # Write the DataFrame to a CSV file
        try:
            processed_df_csv.to_csv(csv_output_path, index=False)
        except Exception as e:
            print(f"An error occurred while writing to {csv_output_path}: {e}")
            return

    # Define value ranges and their corresponding output values for each dataset
    bmx_value_ranges = {
        '-999': (-999, -999, -1),
        '0': (0, 0, 1),
        '15': (0, 15, 2),
        '30': (15, 30, 3),
        '45': (30, 45, 4),
        '60': (45, 60, 5),
        '80': (60, 80, 6),
        '90': (80, 90, 7),
        'more90': (90, float('inf'), 8)
    }
    bmx_nom_value_ranges = {
        '-999': (-999, -999, -1),
        '0': (0, 0, 1),
        '15': (0, 15, 2),
        '30': (15, 30, 3),
        '45': (30, 45, 4),
        '60': (45, 60, 5),
        '80': (60, 80, 6),
        '90': (80, 90, 7),
        'more90': (90, float('inf'), 8)
    }
    # Define value ranges and their corresponding output values for each dataset
    hin_value_ranges = {
        '-999': (-999, -999, -1),
        '0': (0, 0, 1),
        '15': (0, 15, 2),
        '30': (15, 30, 3),
        '45': (30, 45, 4),
        '60': (45, 60, 5),
        '80': (60, 80, 6),
        '90': (80, 90, 7),
        'more90': (90, float('inf'), 8)
    }
    hin_nom_value_ranges = {
        '-999': (-999, -999, -1),
        '0': (0, 0, 1),
        '15': (0, 15, 2),
        '30': (15, 30, 3),
        '45': (30, 45, 4),
        '60': (45, 60, 5),
        '80': (60, 80, 6),
        '90': (80, 90, 7),
        'more90': (90, float('inf'), 8)
    }
    cwp_value_ranges = {
        '-999': (-999, -999, -1),
        '0': (0, 0, 1),
        '0.1': (0, 0.1, 2),
        '0.2': (0.1, 0.2, 3),
        '0.3': (0.2, 0.3, 4),
        '0.5': (0.3, 0.5, 5),
        '1.0': (0.5, 1.0, 6),
        'more1.0': (1.0, float('inf'), 7)
    }
    crop_cycle_value_ranges = {
        '-999': (-999, -999, -1),
        '100': (0, 100, 1),
        '125': (100, 125, 2),
        '150': (125, 150, 3),
        '175': (150, 175, 4),
        '200': (175, 200, 5),
        '250': (200, 250, 6),
        'more300': (300, float('inf'), 7)
    }
    yield_value_ranges = {
        '-999': (-999, -999, -1),
        '0': (0, 0, 1),
        '0.25': (0, 0.25, 2),
        '0.50': (0.25, 0.5, 3),
        '1.0': (0.5, 1.0, 4),
        '2.0': (1.0, 2.0, 5),
        '4.0': (2.0, 4.0, 6),
        'more4.0': (4.0, float('inf'), 7)
    }
    RCF_value_ranges = {
        '-999': (-999, -999, -1),
        '10': (0, 10, 1),
        '20': (10, 20, 2),
        '25': (20, 25, 3),
        '33': (25, 33, 4),
        '50': (33, 50, 5),
        'more50': (50, float('inf'), 6)
    }
    sto_stress_value_ranges = {
        '-999': (-999, -999, -1),
        '0': (0, 0, 1),
        '10': (0, 10, 2),
        '20': (10, 20, 3),
        '30': (20, 30, 4),
        '40': (30, 40, 5),
        '50': (40, 50, 6),
        '75': (50, 75, 7),
        'more75': (75, float('inf'), 8)
    }
    leaf_stress_value_ranges = {
        '-999': (-999, -999, -1),
        '0': (0, 0, 1),
        '10': (0, 10, 2),
        '20': (10, 20, 3),
        '30': (20, 30, 4),
        '40': (30, 40, 5),
        '50': (40, 50, 6),
        '75': (50, 75, 7),
        'more75': (75, float('inf'), 8)
    }
    temp_stress_value_ranges = {
        '-999': (-999, -999, -1),
        '0': (0, 0, 1),
        '10': (0, 10, 2),
        '20': (10, 20, 3),
        '30': (20, 30, 4),
        '40': (30, 40, 5),
        '50': (40, 50, 6),
        '75': (50, 75, 7),
        'more75': (75, float('inf'), 8)
    }
    rfl_value_ranges = {
        '-999': (-999, -999, -1),
        '200': (0, 200, 2),
        '300': (200, 300, 3),
        '400': (300, 400, 4),
        '500': (400, 500, 5),
        '600': (500, 600, 6),
        '700': (600, 700, 7),
        'more700': (700, float('inf'), 8)
    }
    yld_obs_value_ranges = {
        '-999': (-999, -999, -1),
        '10': (0, 10, 2),
        '20': (10, 20, 3),
        '30': (20, 30, 4),
        '40': (30, 40, 5),
        '48': (40, 48, 6),
        '49': (48, float('inf'), 7)
    }
    yld_cv_value_ranges = {
        '-999': (-999, -999, -1),
        '20': (0, 20, 1),
        '30': (20, 30, 2),
        '50': (30, 50, 3),
        '70': (50, 70, 4),
        '100': (70, 100, 5),
        '150': (100, 150, 6),
        '151': (150, float('inf'), 7)
    }
    nwp_k_value_ranges = {
        '-999': (-999, -999, -1),
        '0': (0, 0, 1),
        '1': (0, 1, 2),
        '2': (1, 2, 3),
        '5': (2, 5, 4),
        '10': (5, 10, 5),
        '15': (10, 15, 6),
        'more15': (15, float('inf'), 7)
    }
    nwp_zn_value_ranges = {
        '-999': (-999, -999, -1),
        '0': (0, 0, 1),
        '0.005': (0, 0.005, 2),
        '0.01': (0.005, 0.01, 3),
        '0.02': (0.01, 0.02, 4),
        '0.05': (0.02, 0.05, 5),
        '0.1': (0.05, 0.1, 6),
        'more0.1': (0.1, float('inf'), 7)
    }
    nwp_fe_value_ranges = {
        '-999': (-999, -999, -1),
        '0': (0, 0, 1),
        '0.025': (0, 0.025, 2),
        '0.05': (0.025, 0.05, 3),
        '0.075': (0.05, 0.075, 4),
        '0.1': (0.075, 0.1, 5),
        '0.2': (0.1, 0.2, 6),
        'more0.2': (0.2, float('inf'), 7)
    }

    # Process each dataset
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'BMX',
                     f'obstmp_195099_tardry_{planting_date_dir}_bmx_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir, 'BMX', 'BMX.xlsx'),
        os.path.join(base_output_dir, planting_date_dir, 'BMX', 'BMX.csv'),
        bmx_value_ranges,
        'BMX'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'BMX_nom',
                     f'obstmp_195099_tardry_{planting_date_dir}_bmx_nom_014.csv'),
        os.path.join(base_output_dir, planting_date_dir,
                     'BMX_nom', 'BMX_nom.xlsx'),
        os.path.join(base_output_dir, planting_date_dir,
                     'BMX_nom', 'BMX_nom.csv'),
        bmx_nom_value_ranges,
        'BMX_nom'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'HIN',
                     f'obstmp_195099_tardry_{planting_date_dir}_hin_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir, 'HIN', 'hin.xlsx'),
        os.path.join(base_output_dir, planting_date_dir, 'HIN', 'hin.csv'),
        hin_value_ranges,
        'hin'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'HIN_nom',
                     f'obstmp_195099_tardry_{planting_date_dir}_hin_nom_014.csv'),
        os.path.join(base_output_dir, planting_date_dir,
                     'HIN_nom', 'hin_nom.xlsx'),
        os.path.join(base_output_dir, planting_date_dir,
                     'HIN_nom', 'hin_nom.csv'),
        hin_nom_value_ranges,
        'hin_nom'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'CWP',
                     f'obstmp_195099_tardry_{planting_date_dir}_wue_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir, 'CWP', 'cwp.xlsx'),
        os.path.join(base_output_dir, planting_date_dir, 'CWP', 'cwp.csv'),
        cwp_value_ranges,
        'cwp'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'crop_cycle',
                     f'obstmp_195099_tardry_{planting_date_dir}_len_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir,
                     'crop_cycle', 'crop_cycle.xlsx'),
        os.path.join(base_output_dir, planting_date_dir,
                     'crop_cycle', 'crop_cycle.csv'),
        crop_cycle_value_ranges,
        'crop_cycle'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'yield',
                     f'obstmp_195099_tardry_{planting_date_dir}_yld_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir,
                     'yield', 'yield.xlsx'),
        os.path.join(base_output_dir, planting_date_dir, 'yield', 'yield.csv'),
        yield_value_ranges,
        'yield'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'RCF',
                     f'obstmp_195099_tardry_{planting_date_dir}_rcf_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir, 'RCF', 'RCF.xlsx'),
        os.path.join(base_output_dir, planting_date_dir, 'RCF', 'RCF.csv'),
        RCF_value_ranges,
        'RCF'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'sto_stress',
                     f'obstmp_195099_tardry_{planting_date_dir}_sst_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir,
                     'sto_stress', 'sto_stress.xlsx'),
        os.path.join(base_output_dir, planting_date_dir,
                     'sto_stress', 'sto_stress.csv'),
        sto_stress_value_ranges,
        'sto_stress'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'leaf_stress',
                     f'obstmp_195099_tardry_{planting_date_dir}_sle_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir,
                     'leaf_stress', 'leaf_stress.xlsx'),
        os.path.join(base_output_dir, planting_date_dir,
                     'leaf_stress', 'leaf_stress.csv'),
        leaf_stress_value_ranges,
        'leaf_stress'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'temp_stress',
                     f'obstmp_195099_tardry_{planting_date_dir}_ste_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir,
                     'temp_stress', 'temp_stress.xlsx'),
        os.path.join(base_output_dir, planting_date_dir,
                     'temp_stress', 'temp_stress.csv'),
        temp_stress_value_ranges,
        'temp_stress'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'RFL',
                     f'obstmp_195099_tardry_{planting_date_dir}_rfl_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir, 'RFL', 'RFL.xlsx'),
        os.path.join(base_output_dir, planting_date_dir, 'RFL', 'RFL.csv'),
        rfl_value_ranges,
        'RFL'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'yield_obs',
                     f'obstmp_195099_tardry_{planting_date_dir}_yld_obs_014.csv'),
        os.path.join(base_output_dir, planting_date_dir,
                     'yield_obs', 'Yld_obs.xlsx'),
        os.path.join(base_output_dir, planting_date_dir,
                     'yield_obs', 'Yld_obs.csv'),
        yld_obs_value_ranges,
        'yld_obs'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'yield_cv',
                     f'obstmp_195099_tardry_{planting_date_dir}_yld_pcv_014.csv'),
        os.path.join(base_output_dir, planting_date_dir,
                     'yield_cv', 'Yld_cv.xlsx'),
        os.path.join(base_output_dir, planting_date_dir,
                     'yield_cv', 'Yld_cv.csv'),
        yld_cv_value_ranges,
        'yld_cv'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'nwp_k',
                     f'obstmp_195099_tardry_{planting_date_dir}_nwp_k_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir,
                     'nwp_k', 'nwp_k.xlsx'),
        os.path.join(base_output_dir, planting_date_dir, 'nwp_k', 'nwp_k.csv'),
        nwp_k_value_ranges,
        'nwp_k'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'nwp_fe',
                     f'obstmp_195099_tardry_{planting_date_dir}_nwp_fe_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir,
                     'nwp_fe', 'nwp_fe.xlsx'),
        os.path.join(base_output_dir, planting_date_dir,
                     'nwp_fe', 'nwp_fe.csv'),
        nwp_fe_value_ranges,
        'nwp_fe'
    )
    process_data(
        os.path.join(base_output_dir, planting_date_dir, 'nwp_zn',
                     f'obstmp_195099_tardry_{planting_date_dir}_nwp_zn_ave_014.csv'),
        os.path.join(base_output_dir, planting_date_dir,
                     'nwp_zn', 'nwp_zn.xlsx'),
        os.path.join(base_output_dir, planting_date_dir,
                     'nwp_zn', 'nwp_zn.csv'),
        nwp_zn_value_ranges,
        'nwp_zn'
    )

    time.sleep(2.5)  # Simulating process
    end_time_section4 = time.time()
    print(
        f"Section 4 for {planting_date_dir} took {end_time_section4 - start_time_section4:.2f} seconds.")


# --- Main execution ---
if __name__ == "__main__":
    # Define the base input and output directories
    base_input_dir = 'C:/AquaCrop_crop_output/tar_t54_snew_two_layer/tardry/'
    base_output_dir = 'C:/AquaCrop_crop_output/tar_t54_snew_two_layer/tardry/'

    # List of planting date directories
    planting_dates = ['01a010', '01i010', '01j010',
                      '01k010', '01l010']  # add all dates here

    # Create the base output directory if it doesn't exist
    os.makedirs(base_output_dir, exist_ok=True)

    # Process each planting date
    for planting_date in planting_dates:
        process_planting_date(planting_date, base_output_dir)

    print("Processing of all planting dates completed.")
